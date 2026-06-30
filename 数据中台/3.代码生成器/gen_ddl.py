#!/usr/bin/env python3
"""Generate Doris DDL for MIHDB_DWD from the V1.2 Excel design doc."""
import openpyxl
import re
import sys
from pathlib import Path

XLSX = "/Users/zoyoe/美年健康研究院/数据平台/《DWD_体检数据仓库表结构设计_V1.2》.xlsx"
OUT = Path("/Users/zoyoe/codex/dolphinscheduler/ddl-dwd/MIHDB_DWD.sql")
DB = "MIHDB_DWD"
REPLICATION = 2

def norm_type(raw):
    t = str(raw or "").strip().upper()
    t = t.replace("TINYINT(1)", "TINYINT")
    # Doris does not support NVARCHAR; map to VARCHAR. byte length is 3x char for safety.
    m = re.match(r"NVARCHAR\((\d+)\)", t)
    if m:
        n = int(m.group(1))
        t = f"VARCHAR({min(n*3, 65533)})"
    elif t == "NVARCHAR":
        t = "VARCHAR(255)"
    # Map TEXT to STRING (Doris)
    if t == "NTEXT" or t == "TEXT":
        t = "STRING"
    return t

def cn_to_snake(cn):
    """Convert Chinese column name like "代码值(positive_level_code)" -> snake name."""
    m = re.search(r"[（(]\s*([a-z][a-z0-9_]+)\s*[）)]", str(cn or ""), re.I)
    if m:
        return m.group(1).lower()
    fallback = str(cn or "").strip()
    fallback = re.sub(r"[^a-z0-9_]", "", fallback, flags=re.I)
    return fallback or "col"

def parse_fact(ws):
    header_row = None
    for r in range(1, min(8, ws.max_row)+1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if row and str(row[0]).strip() == "序号":
            header_row = r
            title_row = r-1
            break
    if not header_row:
        return None, None, []
    title = str(list(ws.iter_rows(min_row=title_row, max_row=title_row, values_only=True))[0][0] or "")
    m = re.search(r"[（(]\s*(dwd_[a-z0-9_]+)\s*[）)]", title, re.I)
    if not m:
        m = re.search(r"(dwd_[a-z0-9_]+)", title, re.I)
    tbl = m.group(1).lower() if m else None
    cols = []
    for r in range(header_row+1, ws.max_row+1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row[0]:
            break
        seq, cn, en, typ, pk, nn, default, *_ = row + ('',)*(9-len(row))
        if not en:
            continue
        cols.append({
            "name": str(en).strip(),
            "cn": str(cn or "").strip(),
            "type": norm_type(typ),
            "pk": str(pk or "").strip().upper() == "Y",
            "nn": str(nn or "").strip().upper() == "Y",
            "default": "" if default is None else str(default).strip(),
        })
    return tbl, title, cols

def parse_dim(ws):
    # Read first 10 lines to detect format
    raw_rows = []
    for r in range(1, min(ws.max_row, 11)+1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if any(str(c).strip() for c in row if c is not None):
            raw_rows.append([str(c or "").strip() for c in row])
    if len(raw_rows) < 3:
        return None, None, []
    title = raw_rows[0][0] if raw_rows else ""
    line_cn = raw_rows[1]
    # Type C: column-definition table (e.g. dim_date). line 2 = ["字段名", "类型", "说明"].
    if len(line_cn) >= 3 and line_cn[0] == "字段名" and line_cn[1] == "类型" and line_cn[2] == "说明":
        m = re.search(r"[（(]\s*(dim_[a-z0-9_]+)\s*[）)]", title, re.I)
        if not m:
            m = re.search(r"(dim_[a-z0-9_]+)", title, re.I)
        tbl = m.group(1).lower() if m else None
        cols = []
        first = True
        for r in range(3, ws.max_row + 1):
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            name = str(row[0] or "").strip() if len(row) > 0 else ""
            typ = str(row[1] or "").strip() if len(row) > 1 else ""
            comment = str(row[2] or "").strip() if len(row) > 2 else ""
            if not name or not typ:
                break
            cols.append({
                "name": name.lower(),
                "cn": comment,
                "type": norm_type(typ),
                "pk": first,
                "nn": True,
                "default": ""
            })
            first = False
        return tbl, title, cols

    line_x = raw_rows[2]  # may be eng names (type B) or first data (type A)

    # decide: is line_x data or English header?
    has_chinese_x = any(re.search(r"[\u4e00-\u9fa5]", c) for c in line_x if c.strip())
    if has_chinese_x:
        # Type A: line 2 is cn header, line 3 is first data row
        en_names = [cn_to_snake(c) for c in line_cn]
        first_data_row = 3
    else:
        # Type B: line 2 is cn header, line 3 is eng header
        # Some sheets have a typo in row 3 and a corrected row 4 (also eng); prefer row 4 in that case
        line_x2 = raw_rows[3] if len(raw_rows) > 3 else []
        is_row4_eng = bool(line_x2) and not any(re.search(r"[\u4e00-\u9fa5]", c) for c in line_x2 if c.strip())
        is_row4_data_only = bool(line_x2) and not any("_" in c for c in line_x2 if c)
        if line_x2 and is_row4_eng and not is_row4_data_only:
            # row 4 is also an English header row (typo correction) -> use row 4 as canonical names
            chosen = line_x2
            first_data_row = 5
        else:
            chosen = line_x
            first_data_row = 4
        en_names = [re.sub(r"[^a-zA-Z0-9_]", "", c).lower() for c in chosen]
    m = re.search(r"[（(]\s*(dim_[a-z0-9_]+)\s*[）)]", title, re.I)
    if not m:
        m = re.search(r"(dim_[a-z0-9_]+)", title, re.I)
    tbl = m.group(1).lower() if m else None

    cols = []
    for i, (cn, en) in enumerate(zip(line_cn, en_names)):
        if not en:
            en = cn_to_snake(cn)
        if not en:
            continue
        cols.append({
            "name": en.lower(),
            "cn": cn,
            "type": "VARCHAR(64)",
            "pk": False,
            "nn": False,
            "default": ""
        })
    # PK override for known hierarchical/granular dim tables
    PK_OVERRIDE = {
        "dim_lab_item": "lab_item_code",
        "dim_specimen_item": "specimen_item_code",
        "dim_exam_item": "exam_item_code",
        "dim_age": "age",
    }
    pk_name = PK_OVERRIDE.get(tbl, cols[0]["name"] if cols else None)
    for c in cols:
        if c["name"] == pk_name:
            c["pk"] = True
            c["nn"] = True
    for c in cols:
        if c["name"] == "is_effective":
            c["type"] = "VARCHAR(1)"
        if "desc" in c["name"]:
            c["type"] = "VARCHAR(500)"
        if c["name"].endswith("name"):
            c["type"] = "VARCHAR(200)"
    # 2026-06-24 mobile 是加密密文, 各源最长 44 字符 (2023 base64), 统一扩到 128
    for c in cols:
        if c["name"] == "mobile":
            c["type"] = "VARCHAR(128)"
    # 2026-06-24 external_inspection_code 由 VARCHAR(1) 扩到 VARCHAR(2), 与 dim 字典 01/02/99 对齐
    for c in cols:
        if c["name"] == "external_inspection_code":
            c["type"] = "VARCHAR(2)"
    return tbl, title, cols

def render_field(c, all_pks):
    parts = ["`%s`" % c["name"], c["type"]]
    if c["nn"] or c["name"] in all_pks:
        parts.append("NOT NULL")
    else:
        parts.append("NULL")
    if c["default"]:
        parts.append("DEFAULT '%s'" % c["default"])
    if c["cn"]:
        parts.append("COMMENT '%s'" % c["cn"].replace("'", ""))
    return " ".join(parts)

def render_table(db, tbl, title, cols, replication):
    if not cols:
        return f"-- skipped: {tbl} (no columns)\n"
    pks = [c["name"] for c in cols if c["pk"]]
    if not pks:
        pks = [cols[0]["name"]]
    pk_set = set(pks)
    pk_cols = [c for c in cols if c["name"] in pk_set]
    other_cols = [c for c in cols if c["name"] not in pk_set]
    ordered = pk_cols + other_cols
    lines = [f"DROP TABLE IF EXISTS `{db}`.`{tbl}`;",
             f"CREATE TABLE `{db}`.`{tbl}` ("]
    field_lines = [render_field(c, pk_set) for c in ordered]
    lines.append(",\n".join("  " + f for f in field_lines))
    pk_str = ", ".join(f"`{c}`" for c in pks)
    lines.append(f") ENGINE=OLAP")
    lines.append(f"UNIQUE KEY({pk_str})")
    lines.append(f"COMMENT '{title}'")
    bucket = 32 if not tbl.startswith("dim_") else 4
    lines.append(f"DISTRIBUTED BY HASH({pk_str}) BUCKETS {bucket}")
    props = [f'"replication_num" = "{replication}"',
             '"enable_unique_key_merge_on_write" = "true"',
             '"compression" = "ZSTD"']
    lines.append("PROPERTIES (\n  " + ",\n  ".join(props) + "\n);")
    return "\n".join(lines) + "\n"

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    fact_sheets = ['DWD_人员信息','DWD_体检登记','DWD_实验室检验结果','DWD_体格、影像与功能检查',
                   'DWD_问卷信息','DWD_随访信息','DWD_计算指标','DWD_明细标签']
    out_lines = [
        f"-- Auto-generated from {Path(XLSX).name}",
        f"-- Database: {DB}, replication={REPLICATION}",
        "",
        f"CREATE DATABASE IF NOT EXISTS `{DB}`;",
        f"USE `{DB}`;",
        "",
        "-- ==================== FACT TABLES ====================",
        ""
    ]
    for sheet in fact_sheets:
        ws = wb[sheet]
        tbl, title, cols = parse_fact(ws)
        if not tbl:
            out_lines.append(f"-- ! sheet {sheet}: cannot resolve table name")
            continue
        out_lines.append(f"-- ----- {sheet}: {tbl} -----")
        out_lines.append(render_table(DB, tbl, title, cols, REPLICATION))
        out_lines.append("")

    out_lines.append("-- ==================== DIM TABLES ====================")
    out_lines.append("")
    for sheet in wb.sheetnames:
        if not sheet.startswith("维度表_"): continue
        ws = wb[sheet]
        tbl, title, cols = parse_dim(ws)
        if not tbl:
            out_lines.append(f"-- ! sheet {sheet}: cannot resolve table name")
            continue
        out_lines.append(f"-- ----- {sheet}: {tbl} -----")
        out_lines.append(render_table(DB, tbl, title, cols, REPLICATION))
        out_lines.append("")

    OUT.write_text("\n".join(out_lines), encoding="utf-8")
    print(f"wrote {OUT}")

if __name__ == "__main__":
    main()
