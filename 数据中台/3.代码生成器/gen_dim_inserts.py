#!/usr/bin/env python3
"""Generate INSERT SQL for dimension table dict data from Excel."""
import openpyxl
import re
from pathlib import Path

XLSX = "/Users/zoyoe/美年健康研究院/数据平台/《DWD_体检数据仓库表结构设计_V1.2》.xlsx"
OUT = Path("/Users/zoyoe/codex/dolphinscheduler/ddl-dwd/MIHDB_DWD_dim_data.sql")
DB = "MIHDB_DWD"

def esc(s):
    if s is None:
        return "NULL"
    s = str(s).strip()
    if not s:
        return "NULL"
    # Doris requires backslash escape in strings
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"

def parse(sheet_name, ws):
    """Return (table_name, header_row_num, column_names, data_rows)"""
    raw_rows = []
    for r in range(1, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        raw_rows.append([("" if c is None else str(c)).strip() for c in row])
    title = raw_rows[0][0]
    m = re.search(r"[（(]\s*(dim_[a-z0-9_]+)\s*[）)]", title, re.I)
    if not m:
        m = re.search(r"(dim_[a-z0-9_]+)", title, re.I)
    tbl = m.group(1).lower() if m else None
    if not tbl:
        raise ValueError(f"cannot resolve table name for {sheet_name}")
    # dim_date is column-def table: no data rows to insert
    if tbl == "dim_date":
        return tbl, None, [], []

    # Decide where data starts
    # Row 2 (0-based idx 1) is always Chinese header (may have paren en names)
    row2 = raw_rows[1]
    has_paren_en = any(re.search(r"[（(][a-z0-9_]+[）)]", c, re.I) for c in row2 if c)

    # If row 3 (idx 2) looks like snake_case en names (underscore present) -> skip it (and maybe row 4 if also en)
    header_row_idx = 1
    data_start_idx = 2
    # Skip consecutive snake_case rows: they're English header samples
    while data_start_idx < len(raw_rows) and data_start_idx <= 5:
        row = raw_rows[data_start_idx]
        if not any(c for c in row):
            break
        underscore_count = sum(1 for c in row if c and "_" in c and " " not in c)
        if underscore_count >= len(row) * 0.5 and underscore_count >= 2:
            data_start_idx += 1
        else:
            break

    # Extract column names from paren in row 2 (Type A) or snake_case row (Type B)
    en_from_paren = []
    for c in row2:
        m = re.search(r"[（(]\s*([a-z][a-z0-9_]+)\s*[）)]", c, re.I)
        if m:
            en_from_paren.append(m.group(1).lower())
    if len(en_from_paren) == len(row2):
        col_names = en_from_paren
    else:
        # fallback: take snake_case row that we skipped (if we skipped one)
        col_names = []
        for idx in range(data_start_idx-1, 1, -1):
            row = raw_rows[idx]
            if any("_" in c for c in row if c):
                col_names = [re.sub(r"[^a-z0-9_]", "", c.strip().lower()) for c in row]
                break
        if not col_names:
            raise ValueError(f"cannot resolve column names for {sheet_name}")
    # remove empty col names
    col_names = [c for c in col_names if c]

    # extract data
    data_rows = []
    for idx in range(data_start_idx, len(raw_rows)):
        row = raw_rows[idx]
        if not any(c for c in row):
            continue
        # only take len(col_names) cols
        data_rows.append(row[:len(col_names)])
    return tbl, header_row_idx, col_names, data_rows

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    lines = [f"-- Auto-generated dim data inserts from {Path(XLSX).name}",
             f"USE `{DB}`;",
             ""]
    total_rows = 0
    for sheet_name in sorted(wb.sheetnames):
        if not sheet_name.startswith("维度表_"):
            continue
        ws = wb[sheet_name]
        try:
            tbl, header_row_idx, cols, data = parse(sheet_name, ws)
        except Exception as e:
            lines.append(f"-- ! {sheet_name}: {e}")
            continue
        if tbl == "dim_date" or not data:
            lines.append(f"-- {sheet_name}: no data")
            continue
        lines.append(f"-- {sheet_name}: {len(data)} rows")
        lines.append(f"INSERT INTO `{tbl}` ({','.join(f'`{c}`' for c in cols)}) VALUES")
        val_lines = []
        for d in data:
            vals = ",".join(esc(v) for v in d)
            val_lines.append(f"  ({vals})")
        lines.append(",\n".join(val_lines) + ";")
        lines.append("")
        total_rows += len(data)
    lines.append(f"-- total dimension records: {total_rows}")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUT} ({total_rows} rows across 17 dimension tables)")

if __name__ == "__main__":
    main()
